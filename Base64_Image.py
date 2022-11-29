import  requests
import base64
import os
import requests
import  json
from openpyxl.drawing.image import Image
from openpyxl import Workbook,load_workbook
from PIL import Image,ImageEnhance
import  openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import  get_column_letter
from PIL import ImageFont,ImageDraw,ImageColor
# with open("test/20210809080360240.jpg", 'rb') as f:
#     a = base64.b64encode(f.read())
# print(a)

class ImageBae64():
    def getfiles(self):
        """
        获取文件夹内所有文件路径
        """
        import os
        file = r'D:\桌面的缓存\20220608\测试资料\集团\摄像头画面质检\图片\白天_过曝'
        self.a = []
        self.b = []
        self.c = []
        for root, dirs, files in os.walk(file):
            for file in files:
                path = os.path.join(root, file)
                self.file =file
                self.path = path
                # print("文件是：",file)
                # print("文件地址：",path)
                """
                依次读取文件并转换为base64
                """
                """获取文件的名称"""
                filesname = os.path.basename(path)
                with open(path, 'rb') as f:
                    image_data = base64.b64encode(f.read()).decode()
                    # print(type(a))
                # print(a)

                #容器调用地址
                url = "http://127.0.0.1:32375/VQD/check"
                #北向调用地址
                # url = "http://101.91.213.165:9007/ivmapi/v3/index"
                # types = "blur,brightness,colordeviation,snownise,stripe"
                #容器传参
                types = "blur"
                #北向传参
                # analysisType=[41001]
                payload = json.dumps({  "code": "47302","rsaKey": "67da12af7900424fab","sign": "签名字符串",'type':analysisType,'source':image_data})
                headers = {
                    'Content-Type': "application/json",
                    'cache-control': "no-cache",
                    'Postman-Token': "4ec0e459-88a3-4a92-ab32-7cf4c978ef59"
                    }

                response = requests.request(method="POST", url=url, data=payload, headers=headers)
                print(response.text)
                self.nomol =response.json()["data"]["blur"]["is_abnormal"]
                self.data =response.json()["data"]["blur"]
                print(filesname, self.nomol)
                self.a.append(self.nomol)
                self.b.append(self.data)
                self.c.append(self.file)
        print(self.a)
        print(self.b)
        print(self.c)


        wb = Workbook()
        ws = wb.active
        ws['A1'] = "图片名"
        ws['B1'] = "告警情况"
        ws['C1'] = "具体返回内容"
        for (i, row) in enumerate(self.c):
            print("b里面的值是：", self.c[0])
            ws['A{0}'.format(i + 2)] = str(row)
            print(i, type(row))



        for (i, row) in enumerate(self.a):
            # ws['A{0}'.format(i + 2)] = row[0]
            ws['B{0}'.format(i + 2)] = row

        for (i, row) in enumerate(self.b):
            # print("b里面的值是：",self.b[0])
            ws['C{0}'.format(i + 2)] = str(row)
            # print(i, type(row))

        wb.save("D:/桌面的缓存/20220608/测试资料/集团/摄像头画面质检/图片/模糊2.xlsx")


#
# class ResToexcel():
#     def export_excel(self):
#         # 循环写入excel
#         wb = Workbook()
#         ws = wb.active
#         ws['A1'] = "图片名"
#         ws['B1'] = "告警情况"
#         ws['C1'] = "具体返回内容"
#
#         for (i,row) in enumerate(self.a):
#             # 利用枚举的方法循环读取excel中的数据
#             # i代表下标0,1,2,3，row代表每次循环到的结果(2011, 690, 662)
#             # 有几行sql数据就代表i，row代表单行读取的数据（元组形式）
#             # print(i, row)
#             # ws['A1'] = row(0)
#             # ws['B1'] = row(1)
#             # ws['C1'] = row(2)
#             # 将sql数据写入excel方法一：
#             ws['A{0}'.format(i+2)] = row[0]
#             ws['B{0}'.format(i + 2)] = row[1]
#             # ws['C{0}'.format(i + 2)] = row[2]
#             # 将sql数据写入excel方法二：
#             (ws['A{0}'.format(i + 2)], ws['B{0}'.format(i + 2)],
#              ws['C{0}'.format(i + 2)]) = row
#
#         wb.save("./data_from_mysql.xlsx")
#         return self.export_excel

#
# class ImageToexcel():
#
#     '''
#
#     色值转换：
#
#     从图片读取的像素块色值是 RGB 值,
#
#     RGB 和十六进制色值转换。
#
#     '''
#
#     def rgb_to_hex(self,rgb):
#
#         rgb = rgb.split(',')
#
#         color = ''
#
#         # 循环遍历
#
#         for i in rgb:
#             num = int(i)
#
#             color += str(hex(num))[-2:].replace('x', '0').upper()
#
#         return color
#
#     '''
#
#     图片转换：
#
#     逐行读取图片中的RGB色值，再将RGB色值转换十六进制，填充到excel中
#
#     '''
#
#     def img_to_excel(self,img_path, excel_path):
#
#         # 读取源图片
#
#         img_src = Image.open(img_path)
#
#         # 设置图片宽高
#
#         img_width = img_src.size[0]
#
#         img_hight = img_src.size[1]
#
#         # 图片加载
#
#         str_strlist = img_src.load()
#
#         # 获取当前的excel文件
#
#         wb = openpyxl.Workbook()
#
#         # 保存文件
#
#         wb.save(excel_path)
#
#         # 打开excel_path 下的excel文件，并写入信息
#
#         wb = openpyxl.load_workbook(excel_path)
#
#         cell_width, cell_height = 1.0, 1.0
#
#         # 设置excel的写入页
#
#         sheet = wb['Sheet']
#
#         # 循环图片的高与宽，并存入
#
#         for w in range(img_width):
#
#             for h in range(img_hight):
#                 data = str_strlist[w, h]
#
#                 color = str(data).replace("(", "").replace(")", "")
#
#                 color = self.rgb_to_hex(color)
#
#                 # 设置填充颜色为color
#
#                 fille = PatternFill("solid", fgColor=color)
#
#                 sheet.cell(h + 1, w + 1).fill = fille
#
#         # 循环遍历row,让其全部写入
#
#         for i in range(1, sheet.max_row + 1):
#             sheet.row_dimensions[i].height = cell_height
#
#         # 循环遍历column，让其全部写入
#
#         for i in range(1, sheet.max_column + 1):
#             sheet.column_dimensions[get_column_letter(i)].width = cell_width
#
#         # 保存文件
#
#         wb.save(excel_path)
#
#         # 关闭
#
#         img_src.close()


if __name__ =='__main__':
    img =ImageBae64()
    img.getfiles()
    # res = ResToexcel()
    # res.export_excel()






